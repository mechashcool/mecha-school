"""
Attendance report Excel export using openpyxl.
"""
from io import BytesIO
from datetime import datetime


_STATUS_AR = {'present': 'حاضر', 'absent': 'غائب', 'late': 'متأخر'}


def generate_attendance_excel(rows, report_type='detail', date_from='', date_to='',
                               school=None) -> bytes:
    """
    Build an .xlsx attendance report and return it as bytes.

    rows: list of dicts with keys: student, present, absent, late, checkout, details
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return b''

    wb = Workbook()
    ws = wb.active

    # ── Styles ────────────────────────────────────────────────────────────────
    NAVY = 'FF1A3A5C'
    NAVY_FONT = Font(name='Arial', bold=True, color='FFFFFFFF', size=11)
    HEADER_FILL = PatternFill('solid', fgColor=NAVY)
    ALT_FILL    = PatternFill('solid', fgColor='FFF0F4F8')
    BOLD_FONT   = Font(name='Arial', bold=True, size=10)
    NORMAL_FONT = Font(name='Arial', size=10)
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center',
                              wrap_text=True, readingOrder=2)
    RIGHT_ALIGN  = Alignment(horizontal='right', vertical='center',
                              wrap_text=True, readingOrder=2)
    THIN = Border(
        left=Side(style='thin', color='FFCCCCCC'),
        right=Side(style='thin', color='FFCCCCCC'),
        top=Side(style='thin', color='FFCCCCCC'),
        bottom=Side(style='thin', color='FFCCCCCC'),
    )

    def hdr(cell, text):
        cell.value     = text
        cell.font      = NAVY_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border    = THIN

    def val(cell, text, alt=False):
        cell.value     = text
        cell.font      = NORMAL_FONT
        cell.fill      = ALT_FILL if alt else PatternFill()
        cell.alignment = RIGHT_ALIGN
        cell.border    = THIN

    # ── Sheet title ───────────────────────────────────────────────────────────
    school_name = ''
    if school:
        school_name = getattr(school, 'school_name_ar', '') or getattr(school, 'school_name', '')

    type_labels = {
        'detail':  'تقرير تفصيلي عام',
        'grade':   'تقرير حسب الصف',
        'section': 'تقرير حسب الشعبة',
        'student': 'تقرير حسب الطالب',
        'shift':   'تقرير حسب الشفت',
    }
    ws.title = 'تقرير الحضور'

    row_idx = 1
    ws.cell(row_idx, 1, school_name or 'Mecha School').font = Font(name='Arial', bold=True, size=13)
    row_idx += 1
    ws.cell(row_idx, 1, type_labels.get(report_type, 'تقرير الحضور')).font = BOLD_FONT
    row_idx += 1
    ws.cell(row_idx, 1, f'الفترة: {date_from} — {date_to}').font = NORMAL_FONT
    row_idx += 2

    # ── Summary block ─────────────────────────────────────────────────────────
    total_present = sum(r['present'] for r in rows)
    total_absent  = sum(r['absent']  for r in rows)
    total_late    = sum(r['late']    for r in rows)
    grand_total   = total_present + total_absent + total_late
    grand_pct     = round((total_present + total_late) / grand_total * 100, 1) if grand_total > 0 else 0

    sum_headers = ['عدد الطلاب', 'حاضر', 'متأخر', 'غائب', 'إجمالي السجلات', 'نسبة الحضور']
    for ci, h in enumerate(sum_headers, 1):
        hdr(ws.cell(row_idx, ci), h)
    row_idx += 1
    for ci, v in enumerate([len(rows), total_present, total_late, total_absent, grand_total, f'{grand_pct}%'], 1):
        val(ws.cell(row_idx, ci), v)
    row_idx += 2

    # ── Main table headers ────────────────────────────────────────────────────
    col_headers = ['#', 'اسم الطالب', 'الرقم', 'الصف', 'الشعبة', 'الشفت',
                   'حاضر', 'متأخر', 'غائب', 'الإجمالي', 'نسبة الحضور']
    for ci, h in enumerate(col_headers, 1):
        hdr(ws.cell(row_idx, ci), h)
    row_idx += 1

    # ── Data rows ─────────────────────────────────────────────────────────────
    for i, row in enumerate(rows, 1):
        s       = row['student']
        grade_n = s.section.grade.name if s.section and s.section.grade else '—'
        sec_n   = s.section.name       if s.section else '—'

        # Effective shift name
        shift_n = '—'
        if s.section:
            if s.section.shift_id and hasattr(s.section, 'shift') and s.section.shift:
                shift_n = s.section.shift.name
            elif (s.section.shift_id is None
                  and s.section.grade
                  and s.section.grade.shift_id
                  and hasattr(s.section.grade, 'shift')
                  and s.section.grade.shift):
                shift_n = s.section.grade.shift.name

        total   = row['present'] + row['absent'] + row['late']
        pct_val = round((row['present'] + row['late']) / total * 100, 1) if total else 0
        alt     = (i % 2 == 0)

        row_vals = [i, s.full_name, s.student_id, grade_n, sec_n, shift_n,
                    row['present'], row['late'], row['absent'], total, f'{pct_val}%']
        for ci, v in enumerate(row_vals, 1):
            val(ws.cell(row_idx, ci), v, alt=alt)
        row_idx += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [5, 30, 14, 22, 12, 16, 8, 8, 8, 10, 14]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Detail sheet (individual attendance records) ──────────────────────────
    ws2 = wb.create_sheet('سجل تفصيلي')
    detail_headers = ['#', 'اسم الطالب', 'الصف', 'الشعبة', 'التاريخ',
                      'وقت الحضور', 'وقت الانصراف', 'الحالة', 'المصدر', 'ملاحظات']
    for ci, h in enumerate(detail_headers, 1):
        hdr(ws2.cell(1, ci), h)

    detail_row = 2
    for i, row in enumerate(rows, 1):
        s       = row['student']
        grade_n = s.section.grade.name if s.section and s.section.grade else '—'
        sec_n   = s.section.name       if s.section else '—'
        for att in row.get('details', []):
            alt = (detail_row % 2 == 0)
            status_ar = _STATUS_AR.get(att.status, att.status or '—')
            row_vals = [
                i,
                s.full_name,
                grade_n,
                sec_n,
                att.date.strftime('%Y-%m-%d') if att.date else '—',
                att.check_in.strftime('%H:%M')  if att.check_in  else '—',
                att.check_out.strftime('%H:%M') if att.check_out else '—',
                status_ar,
                att.source or '—',
                att.notes or '',
            ]
            for ci, v in enumerate(row_vals, 1):
                val(ws2.cell(detail_row, ci), v, alt=alt)
            detail_row += 1

    # Column widths for detail sheet
    detail_widths = [5, 28, 20, 12, 13, 12, 12, 10, 10, 22]
    for ci, w in enumerate(detail_widths, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # ── Generated timestamp in both sheets ────────────────────────────────────
    for sheet in [ws, ws2]:
        sheet.cell(row_idx + 2 if sheet is ws else detail_row + 2, 1,
                   f'تم الإنشاء: {datetime.utcnow().strftime("%Y-%m-%d %H:%M")} UTC').font = Font(
            name='Arial', size=8, color='FF999999')

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
