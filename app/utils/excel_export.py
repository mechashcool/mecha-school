"""
Al-Muhandis – Excel Export Utility
Generates .xlsx files using openpyxl.
Usage: from app.utils.excel_export import export_students, export_salary_month
"""
from io import BytesIO
from datetime import datetime


# ─── Shared style helpers ────────────────────────────────────────────────────

def _wb():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                      Border, Side, numbers)
        return Workbook()
    except ImportError:
        return None


def _header_style():
    from openpyxl.styles import Font, PatternFill, Alignment
    return {
        'font':      Font(bold=True, color='FFFFFF', size=11),
        'fill':      PatternFill('solid', fgColor='1A3A5C'),
        'alignment': Alignment(horizontal='center', vertical='center'),
    }


def _apply(cell, **kwargs):
    for k, v in kwargs.items():
        setattr(cell, k, v)


def _autowidth(ws, min_w=12, max_w=50):
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import MergedCell
    for col_idx, col_cells in enumerate(ws.iter_cols(), start=1):
        max_len = 0
        for cell in col_cells:
            if isinstance(cell, MergedCell) or cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_w, max(min_w, max_len + 2))


# ─── STUDENTS ────────────────────────────────────────────────────────────────

def export_students(students) -> bytes | None:
    """Export active students list to Excel. Returns bytes or None."""
    wb = _wb()
    if not wb:
        return None
    from openpyxl.styles import Font, PatternFill, Alignment

    STUDENT_STATUS_AR = {
        'active': 'فعال', 'inactive': 'غير فعال', 'archived': 'مؤرشف',
    }

    ws = wb.active
    ws.title = 'الطلاب'
    ws.sheet_view.rightToLeft = True

    headers = ['#', 'كود الطالب', 'الاسم الكامل', 'الجنس', 'تاريخ الميلاد',
               'الصف / الشعبة', 'ولي الأمر', 'هاتف ولي الأمر', 'الحالة', 'تاريخ التسجيل']
    hs = _header_style()
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = hs['font']
        cell.fill      = hs['fill']
        cell.alignment = hs['alignment']
    ws.row_dimensions[1].height = 22

    for i, s in enumerate(students, 1):
        section_txt = ''
        if s.section:
            section_txt = f"{s.section.grade.name} / {s.section.name}"
        row = [
            i,
            s.student_id,
            s.full_name,
            'ذكر' if s.gender == 'male' else 'أنثى',
            s.date_of_birth.strftime('%Y-%m-%d') if s.date_of_birth else '',
            section_txt,
            s.guardian_name or '',
            s.guardian_phone or '',
            STUDENT_STATUS_AR.get(s.status, s.status),
            s.enrollment_date.strftime('%Y-%m-%d') if s.enrollment_date else '',
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i+1, column=col, value=val)
            cell.alignment = Alignment(vertical='center')
            if i % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='F0F4F8')

    _autowidth(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── EMPLOYEES ───────────────────────────────────────────────────────────────

def export_employees(employees) -> bytes | None:
    wb = _wb()
    if not wb:
        return None
    from openpyxl.styles import Font, PatternFill, Alignment

    EMP_STATUS_AR = {
        'active': 'فعال', 'inactive': 'غير فعال', 'terminated': 'منتهي',
    }

    ws = wb.active
    ws.title = 'الموظفون'
    ws.sheet_view.rightToLeft = True

    headers = ['#', 'رقم الموظف', 'الاسم الكامل', 'المسمى الوظيفي', 'القسم',
               'الهاتف', 'البريد الإلكتروني', 'الراتب الأساسي (د.ع)', 'تاريخ التعيين', 'الحالة']
    hs = _header_style()
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hs['font']; cell.fill = hs['fill']; cell.alignment = hs['alignment']
    ws.row_dimensions[1].height = 22

    for i, e in enumerate(employees, 1):
        row = [
            i, e.employee_id, e.full_name, e.job_title,
            e.department or '', e.phone or '', e.email or '',
            float(e.base_salary),
            e.hire_date.strftime('%Y-%m-%d') if e.hire_date else '',
            EMP_STATUS_AR.get(e.status, e.status),
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i+1, column=col, value=val)
            cell.alignment = Alignment(vertical='center')
            if i % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='F0F4F8')

    _autowidth(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── SALARY MONTH ────────────────────────────────────────────────────────────

def export_salary_month(records, month: int, year: int) -> bytes | None:
    wb = _wb()
    if not wb:
        return None
    from openpyxl.styles import Font, PatternFill, Alignment

    AR_MONTHS = ['', 'يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو',
                 'يوليو', 'أغسطس', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر']

    ws = wb.active
    ws.title = f'رواتب {AR_MONTHS[month]} {year}'
    ws.sheet_view.rightToLeft = True

    headers = ['#', 'رقم الموظف', 'الاسم الكامل', 'المسمى الوظيفي', 'القسم',
               'الراتب الأساسي', 'البدلات', 'الخصومات', 'الصافي',
               'الحالة', 'تاريخ الصرف']
    hs = _header_style()
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hs['font']; cell.fill = hs['fill']; cell.alignment = hs['alignment']
    ws.row_dimensions[1].height = 22

    status_en = {'draft': 'مسودة', 'pending': 'مسودة', 'approved': 'معتمد',
                 'paid': 'مدفوع', 'cancelled': 'ملغي'}

    total_net = 0
    for i, r in enumerate(records, 1):
        net = float(r.net_salary)
        total_net += net
        emp_code = r.employee.employee_id if r.employee else ''
        row = [
            i, emp_code, r.employee_name, r.job_title, r.department,
            float(r.base_salary), float(r.allowances or 0),
            float(r.deductions or 0), net,
            status_en.get(r.status, r.status),
            r.paid_date.strftime('%Y-%m-%d') if r.paid_date else '',
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i+1, column=col, value=val)
            cell.alignment = Alignment(vertical='center')
            if i % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='F0F4F8')

    # Total row
    tr = len(records) + 2
    ws.cell(row=tr, column=3, value='الإجمالي').font = Font(bold=True)
    total_cell = ws.cell(row=tr, column=9, value=total_net)
    total_cell.font = Font(bold=True, color='1AAB6D')
    total_cell.fill = PatternFill('solid', fgColor='E8F8F0')

    _autowidth(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── ATTENDANCE ───────────────────────────────────────────────────────────────

def export_attendance(rows, section_name: str, start: str, end: str) -> bytes | None:
    wb = _wb()
    if not wb:
        return None
    from openpyxl.styles import Font, PatternFill, Alignment

    # ── Sheet 1: summary ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'ملخص'
    ws.sheet_view.rightToLeft = True
    ws.cell(row=1, column=1, value=f'تقرير الحضور — {section_name} — {start} إلى {end}')
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)
    ws.merge_cells('A1:G1')

    headers = ['#', 'كود الطالب', 'الاسم الكامل', 'حاضر', 'غائب', 'متأخر', 'نسبة الحضور %']
    hs = _header_style()
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = hs['font']; cell.fill = hs['fill']; cell.alignment = hs['alignment']

    for i, r in enumerate(rows, 1):
        row_data = [
            i, r['student'].student_id, r['student'].full_name,
            r['present'], r['absent'], r['late'], r['rate']
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i+2, column=col, value=val)
            cell.alignment = Alignment(vertical='center')
            if i % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='F0F4F8')
    _autowidth(ws)

    # ── Sheet 2: per-day detail with check_in / check_out ─────────────────
    ws2 = wb.create_sheet(title='التفاصيل')
    ws2.sheet_view.rightToLeft = True
    ws2.cell(row=1, column=1, value=f'تفاصيل الحضور — {section_name} — {start} إلى {end}')
    ws2.cell(row=1, column=1).font = Font(bold=True, size=13)
    ws2.merge_cells('A1:I1')

    detail_headers = ['#', 'كود الطالب', 'الاسم الكامل', 'التاريخ', 'وقت الحضور', 'وقت الانصراف', 'الحالة', 'المصدر', 'ملاحظات']
    for col, h in enumerate(detail_headers, 1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.font = hs['font']; cell.fill = hs['fill']; cell.alignment = hs['alignment']

    STATUS_MAP = {'present': 'حاضر', 'absent': 'غائب', 'late': 'متأخر'}
    detail_row = 3
    for r in rows:
        for att in r.get('details', []):
            ws2.cell(row=detail_row, column=1, value=detail_row - 2)
            ws2.cell(row=detail_row, column=2, value=r['student'].student_id)
            ws2.cell(row=detail_row, column=3, value=r['student'].full_name)
            ws2.cell(row=detail_row, column=4, value=att.date.strftime('%Y-%m-%d'))
            ws2.cell(row=detail_row, column=5, value=att.check_in.strftime('%H:%M')  if att.check_in  else '')
            ws2.cell(row=detail_row, column=6, value=att.check_out.strftime('%H:%M') if att.check_out else '')
            ws2.cell(row=detail_row, column=7, value=STATUS_MAP.get(att.status, att.status))
            ws2.cell(row=detail_row, column=8, value=att.source or '')
            ws2.cell(row=detail_row, column=9, value=att.notes or '')
            if detail_row % 2 == 0:
                for col in range(1, 10):
                    ws2.cell(row=detail_row, column=col).fill = PatternFill('solid', fgColor='F0F4F8')
            detail_row += 1
    _autowidth(ws2)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_fees(records) -> bytes | None:
    wb = _wb()
    if not wb:
        return None
    from openpyxl.styles import Font, PatternFill, Alignment

    ws = wb.active
    ws.title = 'الرسوم'
    ws.sheet_view.rightToLeft = True

    headers = ['#', 'كود الطالب', 'اسم الطالب', 'نوع الرسم', 'المبلغ الإجمالي',
               'الخصم', 'المدفوع', 'المتبقي', 'الأقساط', 'السنة الدراسية']
    hs = _header_style()
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hs['font']; cell.fill = hs['fill']; cell.alignment = hs['alignment']
    ws.row_dimensions[1].height = 22

    for i, rec in enumerate(records, 1):
        row = [
            i,
            rec.student.student_id if rec.student else '',
            rec.student.full_name if rec.student else '',
            rec.fee_type.name if rec.fee_type else '',
            float(rec.total_amount or 0),
            float(rec.discount or 0),
            float(rec.total_paid),
            float(rec.remaining),
            rec.installments.count() if hasattr(rec.installments, 'count') else len(rec.installments),
            rec.academic_year.name if rec.academic_year else '',
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i+1, column=col, value=val)
            cell.alignment = Alignment(vertical='center')
            if i % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='F0F4F8')

    _autowidth(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── FINANCIAL SUMMARY ────────────────────────────────────────────────────────────

def export_financial(revenues, expenses, year: int) -> bytes | None:
    wb = _wb()
    if not wb:
        return None
    from openpyxl.styles import Font, PatternFill, Alignment

    # Sheet 1: Revenues
    ws1 = wb.active
    ws1.title = 'الإيرادات'
    ws1.sheet_view.rightToLeft = True
    headers = ['التاريخ', 'التصنيف', 'الوصف', 'المبلغ (د.ع)']
    hs = _header_style()
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = hs['font']; cell.fill = hs['fill']
    for i, r in enumerate(revenues, 1):
        ws1.cell(row=i+1, column=1, value=r.date.strftime('%Y-%m-%d'))
        ws1.cell(row=i+1, column=2, value=r.category.name)
        ws1.cell(row=i+1, column=3, value=r.description or '')
        ws1.cell(row=i+1, column=4, value=float(r.amount))
    _autowidth(ws1)

    # Sheet 2: Expenses
    ws2 = wb.create_sheet('المصروفات')
    ws2.sheet_view.rightToLeft = True
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = hs['font']; cell.fill = PatternFill('solid', fgColor='B91C1C')
    for i, e in enumerate(expenses, 1):
        ws2.cell(row=i+1, column=1, value=e.date.strftime('%Y-%m-%d'))
        ws2.cell(row=i+1, column=2, value=e.category.name)
        ws2.cell(row=i+1, column=3, value=e.description or '')
        ws2.cell(row=i+1, column=4, value=float(e.amount))
    _autowidth(ws2)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── EXAMS ───────────────────────────────────────────────────────────────────

def export_exams(
    exams,
    subject_report: bool = False,
    student_search: str = '',
) -> bytes | None:
    """Export exam list or per-student subject report to Excel. Returns bytes or None."""
    wb = _wb()
    if not wb:
        return None
    from openpyxl.styles import Font, PatternFill, Alignment

    ws = wb.active
    hs = _header_style()

    if subject_report:
        ws.title = 'تقرير المادة'
        ws.sheet_view.rightToLeft = True
        headers = ['#', 'اسم الطالب', 'كود الطالب', 'اسم الاختبار',
                   'الفئة', 'تاريخ الاختبار', 'الدرجة', 'الدرجة القصوى', 'التقدير', 'الحالة']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hs['font']; cell.fill = hs['fill']; cell.alignment = hs['alignment']
        ws.row_dimensions[1].height = 22

        SUBJ_FILL = PatternFill('solid', fgColor='2B5494')

        # Group exams by subject for section headers
        row_num = 2
        current_subject_id = None
        counter = 0
        for exam in exams:
            subj_id = exam.subject_id
            if subj_id != current_subject_id:
                current_subject_id = subj_id
                subj_name = exam.subject.name if exam.subject else 'مادة غير محددة'
                # Subject name header spanning all columns
                subj_cell = ws.cell(row=row_num, column=1, value=f'المادة: {subj_name}')
                subj_cell.font = Font(bold=True, color='FFFFFF', size=11)
                subj_cell.fill = SUBJ_FILL
                subj_cell.alignment = Alignment(vertical='center')
                ws.merge_cells(
                    start_row=row_num, start_column=1,
                    end_row=row_num, end_column=len(headers)
                )
                ws.row_dimensions[row_num].height = 20
                row_num += 1

            results_q = exam.results.order_by(None)
            if student_search:
                from app.models import ExamResult, Student

                needle = f'%{student_search}%'
                results_q = results_q.join(
                    Student, ExamResult.student_id == Student.id
                ).filter(
                    Student.full_name.ilike(needle) |
                    Student.student_id.ilike(needle)
                )
            for result in results_q.all():
                counter += 1
                row = [
                    counter,
                    result.student.full_name if result.student else '',
                    result.student.student_id if result.student else '',
                    exam.display_name,
                    exam.exam_type.name if exam.exam_type else '-',
                    exam.exam_date.strftime('%Y-%m-%d') if exam.exam_date else '',
                    float(result.marks),
                    float(exam.max_marks),
                    result.grade_letter or '',
                    'ناجح' if result.is_pass else 'راسب',
                ]
                for col, val in enumerate(row, 1):
                    cell = ws.cell(row=row_num, column=col, value=val)
                    cell.alignment = Alignment(vertical='center')
                    if counter % 2 == 0:
                        cell.fill = PatternFill('solid', fgColor='F0F4F8')
                row_num += 1
    else:
        ws.title = 'الاختبارات'
        ws.sheet_view.rightToLeft = True
        headers = ['#', 'اسم الاختبار', 'نوع الاختبار', 'المادة', 'الصف / الشعبة',
                   'السنة الدراسية', 'تاريخ الاختبار', 'الدرجة القصوى', 'درجة النجاح', 'النتائج', 'نسبة النجاح']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hs['font']; cell.fill = hs['fill']; cell.alignment = hs['alignment']
        ws.row_dimensions[1].height = 22

        for i, exam in enumerate(exams, 1):
            section_txt = ''
            if exam.section:
                grade_name = exam.section.grade.name if exam.section.grade else ''
                section_txt = f"{grade_name} / {exam.section.name}" if grade_name else exam.section.name
            total_results = exam.results.count()
            if total_results:
                pass_count = exam.results.filter_by(is_pass=True).count()
                pass_pct = f'{round(pass_count / total_results * 100)}%'
            else:
                pass_pct = '0%'
            row = [
                i,
                exam.display_name,
                exam.exam_type.name if exam.exam_type else '-',
                exam.subject.name if exam.subject else '',
                section_txt,
                exam.academic_year.name if exam.academic_year else '',
                exam.exam_date.strftime('%Y-%m-%d') if exam.exam_date else '',
                float(exam.max_marks),
                float(exam.pass_marks),
                total_results,
                pass_pct,
            ]
            for col, val in enumerate(row, 1):
                cell = ws.cell(row=i + 1, column=col, value=val)
                cell.alignment = Alignment(vertical='center')
                if i % 2 == 0:
                    cell.fill = PatternFill('solid', fgColor='F0F4F8')

    _autowidth(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── GRADEBOOK (PIVOT) ───────────────────────────────────────────────────────

def export_gradebook(exams, rows) -> bytes | None:
    """Export pivot-table gradebook: one row per student, one column per exam."""
    wb = _wb()
    if not wb:
        return None
    from openpyxl.styles import Font, PatternFill, Alignment

    ws = wb.active
    ws.title = 'دفتر الدرجات'
    ws.sheet_view.rightToLeft = True
    hs = _header_style()

    AVG_HEADER = PatternFill('solid', fgColor='1A5C3A')
    ALT_FILL   = PatternFill('solid', fgColor='F0F4F8')

    # ── header row ───────────────────────────────────────────────────────────
    fixed_headers = ['#', 'اسم الطالب', 'كود الطالب']
    exam_headers  = [
        f"{e.display_name}\n({e.exam_date.strftime('%Y-%m-%d')})"
        for e in exams
    ]
    all_headers = fixed_headers + exam_headers + ['المعدل %']

    for col, h in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = hs['font']
        cell.fill      = hs['fill']
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 36

    ws.cell(row=1, column=len(all_headers)).fill = AVG_HEADER

    # ── data rows ────────────────────────────────────────────────────────────
    for i, row in enumerate(rows, 1):
        data_row = i + 1
        alt      = i % 2 == 0
        student  = row['student']

        def _cell(col, val, *, bold=False, color=None, h_align='left', fill=None,
                  _r=data_row, _alt=alt):
            c = ws.cell(row=_r, column=col, value=val)
            c.alignment = Alignment(horizontal=h_align, vertical='center')
            if bold or color:
                c.font = Font(bold=bold, color=color or '000000')
            if fill:
                c.fill = fill
            elif _alt:
                c.fill = ALT_FILL
            return c

        _cell(1, i, h_align='center')
        _cell(2, student.full_name)
        _cell(3, student.student_id, h_align='center')

        for j, (result, exam) in enumerate(zip(row['cells'], exams)):
            col = j + 4
            if result is not None:
                _cell(col, float(result.marks), bold=True,
                      h_align='center')
            else:
                _cell(col, '—', h_align='center')

        avg_col = len(exams) + 4
        avg     = row['avg']
        if avg is not None:
            _cell(avg_col, avg, bold=True, h_align='center')
        else:
            _cell(avg_col, '—', h_align='center')

    _autowidth(ws)
    for col_idx in range(4, len(exams) + 4):
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = max(
            ws.column_dimensions[letter].width, 14
        )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── EMPLOYEE ATTENDANCE ──────────────────────────────────────────────────────

def export_employee_attendance(rows, date_from: str, date_to: str) -> bytes | None:
    """
    Export employee attendance report to Excel.

    rows: list of stat dicts from employee_attendance_helper.get_employees_attendance_summary()
    Each dict has: employee, working_days, present, late, absent, checked_out, rate, daily.
    daily is a list of per-day dicts: date, status, check_in, check_out, source, device, notes.

    Produces two sheets:
      Sheet 1 – Summary  (one row per employee)
      Sheet 2 – Daily Detail (one row per employee-day)
    """
    wb = _wb()
    if not wb:
        return None
    from openpyxl.styles import Font, PatternFill, Alignment

    hs = _header_style()
    STATUS_AR = {'present': 'حاضر', 'absent': 'غائب', 'late': 'متأخر'}
    ALT_FILL = PatternFill('solid', fgColor='F0F4F8')
    ABSENT_FILL = PatternFill('solid', fgColor='FFE0E0')
    LATE_FILL = PatternFill('solid', fgColor='FFF3CD')

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'ملخص الحضور'
    ws.sheet_view.rightToLeft = True

    title = f'تقرير حضور الموظفين — {date_from} إلى {date_to}'
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
    ws.merge_cells('A1:J1')

    headers = ['#', 'اسم الموظف', 'القسم', 'المسمى الوظيفي',
               'أيام العمل', 'حاضر', 'متأخر', 'غائب', 'انصراف', 'نسبة الحضور %']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = hs['font']
        cell.fill = hs['fill']
        cell.alignment = hs['alignment']

    for i, row in enumerate(rows, 1):
        emp = row['employee']
        row_data = [
            i, emp.full_name, emp.department or '', emp.job_title or '',
            row['working_days'], row['present'], row['late'],
            row['absent'], row['checked_out'], row['rate'],
        ]
        fill = ALT_FILL if i % 2 == 0 else None
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 2, column=col, value=val)
            cell.alignment = Alignment(vertical='center')
            if fill:
                cell.fill = fill

    _autowidth(ws)

    # ── Sheet 2: Daily Detail ─────────────────────────────────────────────────
    ws2 = wb.create_sheet(title='تفاصيل يومية')
    ws2.sheet_view.rightToLeft = True

    ws2.cell(row=1, column=1,
             value=f'سجلات الحضور اليومية — {date_from} إلى {date_to}').font = Font(bold=True, size=13)
    ws2.merge_cells('A1:K1')

    detail_headers = [
        '#', 'اسم الموظف', 'القسم', 'المسمى الوظيفي',
        'التاريخ', 'وقت الحضور', 'وقت الانصراف',
        'الحالة', 'المصدر', 'الجهاز', 'ملاحظات',
    ]
    for col, h in enumerate(detail_headers, 1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.font = hs['font']
        cell.fill = hs['fill']
        cell.alignment = hs['alignment']

    detail_row = 3
    for row in rows:
        emp = row['employee']
        for day in row.get('daily', []):
            status = day.get('status', '')
            dev = day.get('device')
            row_data = [
                detail_row - 2,
                emp.full_name, emp.department or '', emp.job_title or '',
                day['date'].strftime('%Y-%m-%d') if day.get('date') else '',
                day['check_in'].strftime('%H:%M') if day.get('check_in') else '',
                day['check_out'].strftime('%H:%M') if day.get('check_out') else '',
                STATUS_AR.get(status, status),
                day.get('source') or '',
                dev.name if dev else '',
                day.get('notes') or '',
            ]
            if status == 'absent':
                fill = ABSENT_FILL
            elif status == 'late':
                fill = LATE_FILL
            elif detail_row % 2 == 0:
                fill = ALT_FILL
            else:
                fill = None

            for col, val in enumerate(row_data, 1):
                cell = ws2.cell(row=detail_row, column=col, value=val)
                cell.alignment = Alignment(vertical='center')
                if fill:
                    cell.fill = fill

            detail_row += 1

    _autowidth(ws2)
    buf2 = BytesIO()
    wb.save(buf2)
    return buf2.getvalue()


def export_single_employee_attendance(emp_row, date_from: str, date_to: str) -> bytes | None:
    """
    Export a single employee's daily attendance to Excel.
    emp_row: one stat dict from calculate_employee_stats().
    """
    wb = _wb()
    if not wb:
        return None
    from openpyxl.styles import Font, PatternFill, Alignment

    hs = _header_style()
    STATUS_AR = {'present': 'حاضر', 'absent': 'غائب', 'late': 'متأخر'}
    ALT_FILL = PatternFill('solid', fgColor='F0F4F8')
    ABSENT_FILL = PatternFill('solid', fgColor='FFE0E0')
    LATE_FILL = PatternFill('solid', fgColor='FFF3CD')

    emp = emp_row['employee']
    ws = wb.active
    ws.title = 'تفاصيل الحضور'
    ws.sheet_view.rightToLeft = True

    title = f'{emp.full_name} — تقرير الحضور — {date_from} إلى {date_to}'
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
    ws.merge_cells('A1:H1')

    # Compact summary block (row 2 = labels, row 3 = values)
    summary_items = [
        ('أيام العمل', emp_row['working_days']),
        ('حاضر',       emp_row['present']),
        ('متأخر',      emp_row['late']),
        ('غائب',       emp_row['absent']),
        ('انصراف',     emp_row['checked_out']),
        ('نسبة الحضور', f"{emp_row['rate']}%"),
    ]
    for col, (label, val) in enumerate(summary_items, 1):
        lbl_cell = ws.cell(row=2, column=col, value=label)
        lbl_cell.font = Font(bold=True, color='FFFFFF')
        lbl_cell.fill = PatternFill('solid', fgColor='1A3A5C')
        lbl_cell.alignment = Alignment(horizontal='center')
        val_cell = ws.cell(row=3, column=col, value=val)
        val_cell.alignment = Alignment(horizontal='center')

    headers = ['#', 'التاريخ', 'وقت الحضور', 'وقت الانصراف',
               'الحالة', 'المصدر', 'الجهاز', 'ملاحظات']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font = hs['font']
        cell.fill = hs['fill']
        cell.alignment = hs['alignment']

    for i, day in enumerate(emp_row.get('daily', []), 1):
        status = day.get('status', '')
        dev = day.get('device')
        row_data = [
            i,
            day['date'].strftime('%Y-%m-%d') if day.get('date') else '',
            day['check_in'].strftime('%H:%M') if day.get('check_in') else '',
            day['check_out'].strftime('%H:%M') if day.get('check_out') else '',
            STATUS_AR.get(status, status),
            day.get('source') or '',
            dev.name if dev else '',
            day.get('notes') or '',
        ]
        if status == 'absent':
            fill = ABSENT_FILL
        elif status == 'late':
            fill = LATE_FILL
        elif i % 2 == 0:
            fill = ALT_FILL
        else:
            fill = None

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 5, column=col, value=val)
            cell.alignment = Alignment(vertical='center')
            if fill:
                cell.fill = fill

    _autowidth(ws)
    buf3 = BytesIO()
    wb.save(buf3)
    return buf3.getvalue()


# ─── PAYROLL LEDGER REPORT (multi-month, filtered) ────────────────────────────

def export_salary_report(records, totals, arabic_months=None,
                         status_labels=None, title='تقرير الرواتب') -> bytes | None:
    """
    Export a filtered, multi-month payroll ledger to Excel.

    records  : list of SalaryRecord rows (already filtered/ordered by the caller)
    totals   : dict from the blueprint's _report_totals (base/allow/deduct/net/paid)
    Returns bytes or None when openpyxl is unavailable.
    """
    wb = _wb()
    if not wb:
        return None
    from openpyxl.styles import Font, PatternFill, Alignment

    arabic_months = arabic_months or ['']*13
    status_labels = status_labels or {}

    ws = wb.active
    ws.title = 'تقرير الرواتب'
    ws.sheet_view.rightToLeft = True

    headers = ['م', 'الموظف', 'الرقم الوظيفي', 'القسم', 'الشهر/السنة',
               'الأساسي', 'البدلات', 'الخصومات', 'الصافي', 'الحالة',
               'تاريخ الصرف', 'أيام الغياب', 'مرات التأخير']
    hs = _header_style()
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hs['font']; cell.fill = hs['fill']; cell.alignment = hs['alignment']
    ws.row_dimensions[1].height = 22

    for i, r in enumerate(records, 1):
        emp_code = r.employee.employee_id if r.employee else ''
        period   = f"{arabic_months[r.month] if r.month < len(arabic_months) else r.month} {r.year}"
        row = [
            i, r.employee_name, emp_code, r.department or '', period,
            float(r.base_salary or 0), float(r.allowances or 0),
            float(r.deductions or 0), float(r.net_salary or 0),
            status_labels.get(r.status, r.status),
            r.paid_date.strftime('%Y-%m-%d') if r.paid_date else '',
            r.absence_days or 0, r.late_count or 0,
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.alignment = Alignment(vertical='center')
            if r.status == 'cancelled':
                cell.fill = PatternFill('solid', fgColor='FBE9E9')
            elif i % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='F0F4F8')

    # Totals row (monetary totals exclude cancelled — see _report_totals)
    tr = len(records) + 2
    ws.cell(row=tr, column=2, value='الإجمالي (عدا الملغية)').font = Font(bold=True)
    for col, key in ((6, 'total_base'), (7, 'total_allow'),
                     (8, 'total_deduct'), (9, 'total_net')):
        c = ws.cell(row=tr, column=col, value=float(totals.get(key, 0)))
        c.font = Font(bold=True)
        c.fill = PatternFill('solid', fgColor='E8F8F0')

    pr = tr + 1
    ws.cell(row=pr, column=2, value='إجمالي المصروف (المدفوع)').font = Font(bold=True)
    pc = ws.cell(row=pr, column=9, value=float(totals.get('total_paid', 0)))
    pc.font = Font(bold=True, color='1A3A5C')
    pc.fill = PatternFill('solid', fgColor='E8F1FB')

    cr = pr + 1
    ws.cell(row=cr, column=2,
            value=(f"عدد السجلات: {totals.get('count', 0)} | "
                   f"مسودة: {totals.get('count_draft', 0)} | "
                   f"معتمد: {totals.get('count_approved', 0)} | "
                   f"مدفوع: {totals.get('count_paid', 0)} | "
                   f"ملغي: {totals.get('count_cancelled', 0)}")).font = Font(size=10)

    _autowidth(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── EMPLOYEE SALARY STATEMENT (ledger) ───────────────────────────────────────

def export_employee_statement(employee, statement, arabic_months=None,
                              year=None) -> bytes | None:
    """
    Export one employee's salary account statement (ledger) to Excel.

    statement : dict from app.services.payroll.employee_statement()
    Returns bytes or None when openpyxl is unavailable.
    """
    wb = _wb()
    if not wb:
        return None
    from openpyxl.styles import Font, PatternFill, Alignment

    ws = wb.active
    ws.title = 'كشف الراتب'
    ws.sheet_view.rightToLeft = True

    # Title block
    ws.cell(row=1, column=1, value='كشف حساب الراتب').font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f'الموظف: {employee.full_name}').font = Font(bold=True)
    info = []
    if employee.job_title:
        info.append(employee.job_title)
    if employee.department:
        info.append(employee.department)
    if year:
        info.append(f'السنة: {year}')
    ws.cell(row=3, column=1, value=' · '.join(info))

    head_row = 5
    headers = ['التاريخ', 'النوع', 'الوصف', 'مستحق (دائن)', 'مصروف (مدين)',
               'الرصيد', 'المرجع']
    hs = _header_style()
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=head_row, column=col, value=h)
        cell.font = hs['font']; cell.fill = hs['fill']; cell.alignment = hs['alignment']
    ws.row_dimensions[head_row].height = 22

    rows = statement.get('rows', [])
    for i, rrow in enumerate(rows, 1):
        data = [
            rrow['date'].strftime('%Y-%m-%d') if rrow.get('date') else '',
            rrow.get('type', ''),
            rrow.get('description', ''),
            float(rrow.get('credit', 0)) or '',
            float(rrow.get('debit', 0)) or '',
            float(rrow.get('balance', 0)),
            rrow.get('ref', ''),
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=head_row + i, column=col, value=val)
            cell.alignment = Alignment(vertical='center')
            if i % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='F0F4F8')

    tr = head_row + len(rows) + 1
    ws.cell(row=tr, column=3, value='الإجمالي').font = Font(bold=True)
    tc = ws.cell(row=tr, column=4, value=float(statement.get('total_credit', 0)))
    tc.font = Font(bold=True, color='1AAB6D')
    td = ws.cell(row=tr, column=5, value=float(statement.get('total_debit', 0)))
    td.font = Font(bold=True, color='1A3A5C')
    bc = ws.cell(row=tr, column=6, value=float(statement.get('balance', 0)))
    bc.font = Font(bold=True)
    bc.fill = PatternFill('solid', fgColor='FFF6E5')

    _autowidth(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
