"""School inventory / warehouse module."""
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from flask import Blueprint, Response, abort, flash, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required
from sqlalchemy import func, or_
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import joinedload

from app.models import (
    db, InventoryCategory, InventoryCount, InventoryItem, InventoryItemStock,
    InventoryMovement, InventoryWarehouse,
)
from app.utils.decorators import (admin_required, permission_required,
                                   any_permission_required,
                                   get_active_year, get_current_school, historical_guard)


inventory_bp = Blueprint('inventory', __name__, template_folder='../../templates/inventory')


CATEGORY_NAMES = ['قرطاسية', 'كتب', 'زي مدرسي', 'ملابس', 'مواد تنظيف', 'أجهزة', 'أثاث', 'أخرى']
DEFAULT_WAREHOUSE_NAME = 'المخزن الرئيسي'
MOVEMENT_TYPES = {'in': 'إدخال', 'out': 'إخراج', 'transfer': 'نقل بين المخازن'}
MOVEMENT_REASONS = [
    'شراء جديد', 'تبرع', 'مرتجع', 'تسليم للطلاب', 'تسليم للمدرسين',
    'صرف للإدارة', 'تلف', 'فقدان', 'استخدام داخلي', 'نقل بين المخازن',
]
COUNT_STATUS = {'match': 'مطابق', 'shortage': 'نقص', 'surplus': 'زيادة'}

# Categories that require a size field (clothing / uniform items).
CLOTHING_CATEGORIES = {'زي مدرسي', 'ملابس', 'uniforms', 'clothing', 'school uniform'}
CLOTHING_SIZES = ['XS', 'S', 'M', 'L', 'XL', 'XXL', 'XXXL',
                  '4', '6', '8', '10', '12', '14', '16']


def _school_year_or_redirect():
    school = get_current_school()
    year = get_active_year(school.id) if school else None
    if not school or not year:
        flash('يرجى اختيار مدرسة وسنة دراسية فعالة قبل استخدام المخازن.', 'danger')
        return None, None, redirect(url_for('admin.dashboard'))
    return school, year, None


def _decimal_value(name, default='0'):
    raw = (request.form.get(name) or default).strip()
    try:
        return Decimal(raw or default)
    except (InvalidOperation, ValueError):
        return Decimal(default)


def _parse_date(name, default=None):
    raw = (request.form.get(name) or '').strip()
    if not raw:
        return default or date.today()
    try:
        return datetime.strptime(raw, '%Y-%m-%d').date()
    except ValueError:
        return default or date.today()


def _ensure_default_categories(school, year):
    existing = {
        c.name for c in InventoryCategory.query
        .filter_by(school_id=school.id, academic_year_id=year.id)
        .all()
    }
    missing = [name for name in CATEGORY_NAMES if name not in existing]
    if missing:
        for name in missing:
            db.session.add(InventoryCategory(
                school_id=school.id,
                academic_year_id=year.id,
                name=name,
            ))
        db.session.commit()


def _categories(school, year):
    return (InventoryCategory.query
            .filter_by(school_id=school.id, academic_year_id=year.id)
            .order_by(InventoryCategory.name)
            .all())


def _clothing_category_ids(categories):
    """Return IDs of categories whose names match the clothing set."""
    lower_names = {n.lower() for n in CLOTHING_CATEGORIES}
    return [c.id for c in categories if c.name.strip().lower() in lower_names]


def _items_query(school, year):
    return (InventoryItem.query
            .options(joinedload(InventoryItem.category))
            .filter_by(school_id=school.id, academic_year_id=year.id))


def _item_or_404(item_id, school, year):
    return _items_query(school, year).filter_by(id=item_id).first_or_404()


def _ensure_default_warehouse(school):
    """Return the school's default warehouse, creating one if none exists yet."""
    warehouse = InventoryWarehouse.query.filter_by(school_id=school.id, is_default=True).first()
    if warehouse:
        return warehouse
    # A warehouse already exists (e.g. created before is_default was set) — promote it.
    warehouse = (InventoryWarehouse.query
                 .filter_by(school_id=school.id)
                 .order_by(InventoryWarehouse.id)
                 .first())
    if warehouse:
        warehouse.is_default = True
        db.session.commit()
        return warehouse
    warehouse = InventoryWarehouse(school_id=school.id, name=DEFAULT_WAREHOUSE_NAME, is_default=True)
    db.session.add(warehouse)
    db.session.commit()
    return warehouse


def _warehouses(school):
    return (InventoryWarehouse.query
            .filter_by(school_id=school.id, is_active=True)
            .order_by(InventoryWarehouse.name)
            .all())


def _all_warehouses(school):
    return (InventoryWarehouse.query
            .filter_by(school_id=school.id)
            .order_by(InventoryWarehouse.name)
            .all())


def _get_or_create_stock(item, warehouse, minimum_quantity=None):
    stock = InventoryItemStock.query.filter_by(item_id=item.id, warehouse_id=warehouse.id).first()
    if not stock:
        stock = InventoryItemStock(
            school_id=item.school_id,
            academic_year_id=item.academic_year_id,
            item_id=item.id,
            warehouse_id=warehouse.id,
            quantity=0,
            minimum_quantity=minimum_quantity if minimum_quantity is not None else (item.minimum_quantity or 0),
        )
        db.session.add(stock)
        db.session.flush()
    return stock


def _recompute_item_total(item):
    """Keep InventoryItem.current_quantity in sync as the sum of its per-warehouse stock."""
    total = (db.session.query(func.coalesce(func.sum(InventoryItemStock.quantity), 0))
             .filter(InventoryItemStock.item_id == item.id)
             .scalar())
    item.current_quantity = total or 0


@inventory_bp.route('/')
@login_required
@any_permission_required('view_inventory', 'manage_inventory')
def index():
    school, year, response = _school_year_or_redirect()
    if response:
        return response
    _ensure_default_categories(school, year)
    _ensure_default_warehouse(school)

    q = request.args.get('q', '').strip()
    category_id = request.args.get('category_id', type=int)
    status = request.args.get('status', 'all')
    warehouse_id = request.args.get('warehouse_id', type=int)
    selected_warehouse = None
    if warehouse_id:
        selected_warehouse = InventoryWarehouse.query.filter_by(id=warehouse_id, school_id=school.id).first()
        if not selected_warehouse:
            warehouse_id = None  # invalid or foreign warehouse id — ignore the filter

    query = _items_query(school, year)
    if q:
        like = f'%{q}%'
        query = query.filter((InventoryItem.name.ilike(like)) | (InventoryItem.item_code.ilike(like)))
    if category_id:
        query = query.filter(InventoryItem.category_id == category_id)
    if warehouse_id:
        query = query.filter(InventoryItem.id.in_(
            db.session.query(InventoryItemStock.item_id).filter(InventoryItemStock.warehouse_id == warehouse_id)
        ))
    if status == 'low':
        query = query.filter(InventoryItem.current_quantity <= InventoryItem.minimum_quantity)

    items = query.order_by(InventoryItem.name).all()
    low_count = sum(1 for item in items if item.is_low_stock)

    # Per-item warehouse breakdown, so search results show every warehouse an
    # item is stocked in (and the specific quantity when filtered to one).
    stocks_by_item = {}
    item_ids = [item.id for item in items]
    if item_ids:
        for stock in (InventoryItemStock.query
                      .options(joinedload(InventoryItemStock.warehouse))
                      .filter(InventoryItemStock.item_id.in_(item_ids))
                      .all()):
            stocks_by_item.setdefault(stock.item_id, []).append(stock)

    return render_template('inventory/index.html',
                           items=items,
                           categories=_categories(school, year),
                           warehouses=_warehouses(school),
                           q=q,
                           category_id=category_id,
                           status=status,
                           warehouse_id=warehouse_id,
                           selected_warehouse=selected_warehouse,
                           stocks_by_item=stocks_by_item,
                           low_count=low_count)


@inventory_bp.route('/items/create', methods=['GET', 'POST'])
@login_required
@historical_guard
@permission_required('manage_inventory')
def create_item():
    school, year, response = _school_year_or_redirect()
    if response:
        return response
    _ensure_default_categories(school, year)
    _ensure_default_warehouse(school)
    categories = _categories(school, year)
    warehouses = _warehouses(school)

    if request.method == 'POST':
        warehouse_id = request.form.get('warehouse_id', type=int)
        warehouse = (InventoryWarehouse.query
                     .filter_by(id=warehouse_id, school_id=school.id, is_active=True)
                     .first()) if warehouse_id else None
        if not warehouse:
            flash('يرجى اختيار مخزن صحيح.', 'danger')
            clothing_ids = _clothing_category_ids(categories)
            return render_template('inventory/item_form.html', item=None, categories=categories,
                                   warehouses=warehouses,
                                   clothing_category_ids=clothing_ids,
                                   clothing_sizes=CLOTHING_SIZES), 400

        item = InventoryItem(
            school_id=school.id,
            academic_year_id=year.id,
        )
        _populate_item(item)
        db.session.add(item)
        db.session.flush()

        stock = _get_or_create_stock(item, warehouse, minimum_quantity=item.minimum_quantity)
        stock.quantity = _decimal_value('current_quantity')
        _recompute_item_total(item)
        db.session.commit()
        flash('تمت إضافة المادة بنجاح.', 'success')
        return redirect(url_for('inventory.index'))

    clothing_ids = _clothing_category_ids(categories)
    return render_template('inventory/item_form.html', item=None, categories=categories,
                           warehouses=warehouses,
                           clothing_category_ids=clothing_ids,
                           clothing_sizes=CLOTHING_SIZES)


@inventory_bp.route('/items/<int:item_id>')
@login_required
@any_permission_required('view_inventory', 'manage_inventory')
def item_detail(item_id):
    school, year, response = _school_year_or_redirect()
    if response:
        return response
    item = _item_or_404(item_id, school, year)
    stocks = (InventoryItemStock.query
              .options(joinedload(InventoryItemStock.warehouse))
              .filter_by(item_id=item.id)
              .join(InventoryWarehouse)
              .order_by(InventoryWarehouse.name)
              .all())
    recent_movements = (InventoryMovement.query
                        .options(joinedload(InventoryMovement.warehouse), joinedload(InventoryMovement.to_warehouse))
                        .filter_by(item_id=item.id)
                        .order_by(InventoryMovement.movement_date.desc(), InventoryMovement.id.desc())
                        .limit(50)
                        .all())
    return render_template('inventory/item_detail.html',
                           item=item, stocks=stocks,
                           movements=recent_movements,
                           movement_types=MOVEMENT_TYPES)


@inventory_bp.route('/items/<int:item_id>/edit', methods=['GET', 'POST'])
@login_required
@historical_guard
@permission_required('manage_inventory')
def edit_item(item_id):
    school, year, response = _school_year_or_redirect()
    if response:
        return response
    item = _item_or_404(item_id, school, year)
    categories = _categories(school, year)
    warehouses = _warehouses(school)

    if request.method == 'POST':
        _populate_item(item)

        # Update existing per-warehouse stock rows submitted with the form.
        stock_ids = request.form.getlist('stock_id', type=int)
        stock_quantities = request.form.getlist('stock_quantity')
        stock_minimums = request.form.getlist('stock_minimum')
        for stock_id, qty_raw, min_raw in zip(stock_ids, stock_quantities, stock_minimums):
            stock = InventoryItemStock.query.filter_by(id=stock_id, item_id=item.id).first()
            if not stock:
                continue
            try:
                stock.quantity = Decimal(qty_raw or '0')
            except InvalidOperation:
                pass
            try:
                stock.minimum_quantity = Decimal(min_raw or '0')
            except InvalidOperation:
                pass

        # Optionally add the item to one more warehouse it isn't stocked in yet.
        new_warehouse_id = request.form.get('new_warehouse_id', type=int)
        if new_warehouse_id:
            new_warehouse = InventoryWarehouse.query.filter_by(
                id=new_warehouse_id, school_id=school.id, is_active=True).first()
            already_assigned = InventoryItemStock.query.filter_by(
                item_id=item.id, warehouse_id=new_warehouse_id).first()
            if not new_warehouse:
                flash('المخزن المحدد للإضافة غير صالح.', 'warning')
            elif already_assigned:
                flash('المادة موجودة بالفعل في هذا المخزن.', 'warning')
            else:
                stock = _get_or_create_stock(item, new_warehouse, minimum_quantity=item.minimum_quantity)
                stock.quantity = _decimal_value('new_quantity')
                stock.minimum_quantity = _decimal_value('new_minimum')

        _recompute_item_total(item)
        db.session.commit()
        flash('تم تحديث المادة بنجاح.', 'success')
        return redirect(url_for('inventory.index'))

    stocks = (InventoryItemStock.query
              .options(joinedload(InventoryItemStock.warehouse))
              .filter_by(item_id=item.id)
              .join(InventoryWarehouse)
              .order_by(InventoryWarehouse.name)
              .all())
    assigned_warehouse_ids = {s.warehouse_id for s in stocks}
    available_warehouses = [w for w in warehouses if w.id not in assigned_warehouse_ids]

    clothing_ids = _clothing_category_ids(categories)
    return render_template('inventory/item_form.html', item=item, categories=categories,
                           warehouses=warehouses, stocks=stocks,
                           available_warehouses=available_warehouses,
                           clothing_category_ids=clothing_ids,
                           clothing_sizes=CLOTHING_SIZES)


def _populate_item(item):
    item.name = request.form.get('name', '').strip()
    item.category_id = request.form.get('category_id', type=int)
    item.item_code = request.form.get('item_code', '').strip() or None
    item.unit = request.form.get('unit', '').strip() or 'قطعة'
    item.size = request.form.get('size', '').strip() or None
    # minimum_quantity here is the item-level default reorder threshold, used
    # to pre-fill new per-warehouse stock rows. current_quantity is no longer
    # set here — it is a derived aggregate (see _recompute_item_total).
    item.minimum_quantity = _decimal_value('minimum_quantity')
    item.purchase_price = _decimal_value('purchase_price') if request.form.get('purchase_price') else None
    item.supplier = request.form.get('supplier', '').strip() or None
    item.notes = request.form.get('notes', '').strip() or None
    item.is_active = bool(request.form.get('is_active', '1'))


@inventory_bp.route('/categories', methods=['GET', 'POST'])
@login_required
@historical_guard
@permission_required('manage_inventory')
def categories():
    school, year, response = _school_year_or_redirect()
    if response:
        return response
    _ensure_default_categories(school, year)

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip() or None
        if not name:
            flash('اسم التصنيف مطلوب.', 'danger')
        elif InventoryCategory.query.filter_by(school_id=school.id, academic_year_id=year.id, name=name).first():
            flash('هذا التصنيف موجود بالفعل.', 'warning')
        else:
            db.session.add(InventoryCategory(
                school_id=school.id,
                academic_year_id=year.id,
                name=name,
                description=description,
            ))
            db.session.commit()
            flash('تمت إضافة التصنيف بنجاح.', 'success')
        return redirect(url_for('inventory.categories'))

    return render_template('inventory/categories.html', categories=_categories(school, year))


@inventory_bp.route('/categories/<int:category_id>/delete', methods=['POST'])
@login_required
@historical_guard
@permission_required('manage_inventory')
def delete_category(category_id):
    school, year, response = _school_year_or_redirect()
    if response:
        return response
    category = InventoryCategory.query.filter_by(
        id=category_id,
        school_id=school.id,
        academic_year_id=year.id,
    ).first()
    if not category:
        flash('التصنيف غير موجود أو تم حذفه مسبقاً.', 'warning')
        return redirect(url_for('inventory.categories'))

    if category.items.count():
        flash('لا يمكن حذف هذا التصنيف لأنه يحتوي على مواد مخزنية.', 'danger')
        return redirect(url_for('inventory.categories'))

    db.session.delete(category)
    db.session.commit()
    flash('تم حذف التصنيف بنجاح.', 'success')
    return redirect(url_for('inventory.categories'))


@inventory_bp.route('/warehouses', methods=['GET', 'POST'])
@login_required
@historical_guard
@permission_required('manage_inventory')
def warehouses():
    school, year, response = _school_year_or_redirect()
    if response:
        return response
    _ensure_default_warehouse(school)

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip() or None
        if not name:
            flash('اسم المخزن مطلوب.', 'danger')
        elif InventoryWarehouse.query.filter_by(school_id=school.id, name=name).first():
            flash('هذا المخزن موجود بالفعل.', 'warning')
        else:
            db.session.add(InventoryWarehouse(school_id=school.id, name=name, description=description))
            db.session.commit()
            flash('تمت إضافة المخزن بنجاح.', 'success')
        return redirect(url_for('inventory.warehouses'))

    return render_template('inventory/warehouses.html', warehouses=_all_warehouses(school))


@inventory_bp.route('/warehouses/<int:warehouse_id>/edit', methods=['POST'])
@login_required
@historical_guard
@permission_required('manage_inventory')
def edit_warehouse(warehouse_id):
    school, year, response = _school_year_or_redirect()
    if response:
        return response
    warehouse = InventoryWarehouse.query.filter_by(id=warehouse_id, school_id=school.id).first()
    if not warehouse:
        flash('المخزن غير موجود.', 'warning')
        return redirect(url_for('inventory.warehouses'))

    name = request.form.get('name', '').strip()
    description = request.form.get('description', '').strip() or None
    if not name:
        flash('اسم المخزن مطلوب.', 'danger')
    elif InventoryWarehouse.query.filter(
            InventoryWarehouse.school_id == school.id,
            InventoryWarehouse.name == name,
            InventoryWarehouse.id != warehouse.id).first():
        flash('هذا الاسم مستخدم بالفعل في مخزن آخر.', 'warning')
    else:
        warehouse.name = name
        warehouse.description = description
        db.session.commit()
        flash('تم تحديث المخزن بنجاح.', 'success')
    return redirect(url_for('inventory.warehouses'))


@inventory_bp.route('/warehouses/<int:warehouse_id>/toggle', methods=['POST'])
@login_required
@historical_guard
@permission_required('manage_inventory')
def toggle_warehouse(warehouse_id):
    school, year, response = _school_year_or_redirect()
    if response:
        return response
    warehouse = InventoryWarehouse.query.filter_by(id=warehouse_id, school_id=school.id).first()
    if not warehouse:
        flash('المخزن غير موجود.', 'warning')
        return redirect(url_for('inventory.warehouses'))
    if warehouse.is_default and warehouse.is_active:
        flash('لا يمكن تعطيل المخزن الافتراضي.', 'danger')
        return redirect(url_for('inventory.warehouses'))

    warehouse.is_active = not warehouse.is_active
    db.session.commit()
    flash('تم تحديث حالة المخزن بنجاح.', 'success')
    return redirect(url_for('inventory.warehouses'))


@inventory_bp.route('/warehouses/<int:warehouse_id>/delete', methods=['POST'])
@login_required
@historical_guard
@permission_required('manage_inventory')
def delete_warehouse(warehouse_id):
    school, year, response = _school_year_or_redirect()
    if response:
        return response
    warehouse = InventoryWarehouse.query.filter_by(id=warehouse_id, school_id=school.id).first()
    if not warehouse:
        flash('المخزن غير موجود.', 'warning')
        return redirect(url_for('inventory.warehouses'))
    if warehouse.is_default:
        flash('لا يمكن حذف المخزن الافتراضي.', 'danger')
        return redirect(url_for('inventory.warehouses'))

    has_stock = InventoryItemStock.query.filter(
        InventoryItemStock.warehouse_id == warehouse.id,
        InventoryItemStock.quantity != 0,
    ).first()
    has_movements = InventoryMovement.query.filter(
        or_(InventoryMovement.warehouse_id == warehouse.id,
            InventoryMovement.to_warehouse_id == warehouse.id),
    ).first()
    has_counts = InventoryCount.query.filter(InventoryCount.warehouse_id == warehouse.id).first()
    if has_stock or has_movements or has_counts:
        flash('لا يمكن حذف هذا المخزن لأنه يحتوي على كميات أو له سجل حركات/جرد. يمكنك تعطيله بدلاً من ذلك.', 'danger')
        return redirect(url_for('inventory.warehouses'))

    # Safe to remove: only zero-quantity, history-free stock rows remain.
    InventoryItemStock.query.filter_by(warehouse_id=warehouse.id).delete(synchronize_session=False)
    db.session.delete(warehouse)
    db.session.commit()
    flash('تم حذف المخزن بنجاح.', 'success')
    return redirect(url_for('inventory.warehouses'))


@inventory_bp.route('/movements')
@login_required
@any_permission_required('view_inventory', 'manage_inventory')
def movements():
    school, year, response = _school_year_or_redirect()
    if response:
        return response
    query = (InventoryMovement.query
             .options(joinedload(InventoryMovement.item), joinedload(InventoryMovement.warehouse),
                      joinedload(InventoryMovement.to_warehouse))
             .filter_by(school_id=school.id, academic_year_id=year.id))
    item_id = request.args.get('item_id', type=int)
    warehouse_id = request.args.get('warehouse_id', type=int)
    if item_id:
        query = query.filter(InventoryMovement.item_id == item_id)
    if warehouse_id:
        query = query.filter(or_(InventoryMovement.warehouse_id == warehouse_id,
                                  InventoryMovement.to_warehouse_id == warehouse_id))
    movements = query.order_by(InventoryMovement.movement_date.desc(), InventoryMovement.id.desc()).all()
    return render_template('inventory/movements.html',
                           movements=movements,
                           movement_types=MOVEMENT_TYPES,
                           items=_items_query(school, year).order_by(InventoryItem.name).all(),
                           warehouses=_warehouses(school),
                           selected_item_id=item_id,
                           selected_warehouse_id=warehouse_id)


@inventory_bp.route('/movements/create', methods=['GET', 'POST'])
@login_required
@historical_guard
@permission_required('manage_inventory')
def create_movement():
    school, year, response = _school_year_or_redirect()
    if response:
        return response
    _ensure_default_warehouse(school)
    items = _items_query(school, year).filter_by(is_active=True).order_by(InventoryItem.name).all()
    warehouses = _warehouses(school)
    selected_item_id = request.args.get('item_id', type=int)

    def _render_error(status=400):
        return render_template('inventory/movement_form.html',
                               items=items,
                               warehouses=warehouses,
                               movement_types=MOVEMENT_TYPES,
                               reasons=MOVEMENT_REASONS,
                               selected_item_id=selected_item_id), status

    if request.method == 'POST':
        item = _item_or_404(request.form.get('item_id', type=int), school, year)
        movement_type = request.form.get('movement_type')
        reason = request.form.get('reason')
        quantity = _decimal_value('quantity')
        warehouse_id = request.form.get('warehouse_id', type=int)
        to_warehouse_id = request.form.get('to_warehouse_id', type=int)

        if movement_type not in MOVEMENT_TYPES or quantity <= 0:
            flash('يرجى إدخال بيانات حركة صحيحة.', 'danger')
            return _render_error()
        if reason not in MOVEMENT_REASONS:
            flash('يرجى إدخال بيانات حركة صحيحة.', 'danger')
            return _render_error()

        warehouse = (InventoryWarehouse.query
                     .filter_by(id=warehouse_id, school_id=school.id, is_active=True)
                     .first()) if warehouse_id else None
        if not warehouse:
            flash('يرجى اختيار مخزن صحيح.', 'danger')
            return _render_error()

        if movement_type == 'transfer':
            to_warehouse = (InventoryWarehouse.query
                            .filter_by(id=to_warehouse_id, school_id=school.id, is_active=True)
                            .first()) if to_warehouse_id else None
            if not to_warehouse or to_warehouse.id == warehouse.id:
                flash('يرجى اختيار مخزن وجهة صحيح ومختلف عن مخزن المصدر.', 'danger')
                return _render_error()
            source_stock = _get_or_create_stock(item, warehouse)
            if source_stock.quantity < quantity:
                flash('الكمية المطلوبة أكبر من الكمية المتوفرة في مخزن المصدر.', 'danger')
                return _render_error()
            dest_stock = _get_or_create_stock(item, to_warehouse)
            source_stock.quantity = source_stock.quantity - quantity
            dest_stock.quantity = dest_stock.quantity + quantity
            movement = InventoryMovement(
                school_id=school.id,
                academic_year_id=year.id,
                item_id=item.id,
                warehouse_id=warehouse.id,
                to_warehouse_id=to_warehouse.id,
                movement_type='transfer',
                reason=reason,
                quantity=quantity,
                movement_date=_parse_date('movement_date'),
                recipient=request.form.get('recipient', '').strip() or None,
                notes=request.form.get('notes', '').strip() or None,
                created_by=current_user.id,
            )
        else:
            stock = _get_or_create_stock(item, warehouse)
            if movement_type == 'out' and stock.quantity < quantity:
                flash('الكمية المطلوبة أكبر من الكمية المتوفرة في هذا المخزن.', 'danger')
                return _render_error()
            stock.quantity = stock.quantity + quantity if movement_type == 'in' else stock.quantity - quantity
            movement = InventoryMovement(
                school_id=school.id,
                academic_year_id=year.id,
                item_id=item.id,
                warehouse_id=warehouse.id,
                movement_type=movement_type,
                reason=reason,
                quantity=quantity,
                movement_date=_parse_date('movement_date'),
                recipient=request.form.get('recipient', '').strip() or None,
                notes=request.form.get('notes', '').strip() or None,
                created_by=current_user.id,
            )

        db.session.add(movement)
        _recompute_item_total(item)
        db.session.commit()
        flash('تم تسجيل حركة المخزن بنجاح.', 'success')
        return redirect(url_for('inventory.movements'))

    return render_template('inventory/movement_form.html',
                           items=items,
                           warehouses=warehouses,
                           movement_types=MOVEMENT_TYPES,
                           reasons=MOVEMENT_REASONS,
                           selected_item_id=selected_item_id)


@inventory_bp.route('/counts')
@login_required
@any_permission_required('view_inventory', 'manage_inventory')
def counts():
    school, year, response = _school_year_or_redirect()
    if response:
        return response
    rows = (InventoryCount.query
            .options(joinedload(InventoryCount.item), joinedload(InventoryCount.counter),
                     joinedload(InventoryCount.warehouse))
            .filter_by(school_id=school.id, academic_year_id=year.id)
            .order_by(InventoryCount.count_date.desc(), InventoryCount.id.desc())
            .all())
    return render_template('inventory/counts.html', counts=rows, status_labels=COUNT_STATUS)


@inventory_bp.route('/counts/create', methods=['GET', 'POST'])
@login_required
@historical_guard
@permission_required('manage_inventory')
def create_count():
    school, year, response = _school_year_or_redirect()
    if response:
        return response
    _ensure_default_warehouse(school)
    items = _items_query(school, year).filter_by(is_active=True).order_by(InventoryItem.name).all()
    warehouses = _warehouses(school)

    if request.method == 'POST':
        try:
            item = _item_or_404(request.form.get('item_id', type=int), school, year)
            warehouse_id = request.form.get('warehouse_id', type=int)
            warehouse = (InventoryWarehouse.query
                        .filter_by(id=warehouse_id, school_id=school.id)
                        .first()) if warehouse_id else None
            if not warehouse:
                flash('يرجى اختيار مخزن صحيح.', 'danger')
                return render_template('inventory/count_form.html', items=items, warehouses=warehouses), 400

            stock = _get_or_create_stock(item, warehouse)
            actual = _decimal_value('actual_quantity')
            diff = actual - stock.quantity
            status = 'match' if diff == 0 else ('surplus' if diff > 0 else 'shortage')
            row = InventoryCount(
                school_id=school.id,
                academic_year_id=year.id,
                item_id=item.id,
                warehouse_id=warehouse.id,
                system_quantity=stock.quantity,
                actual_quantity=actual,
                difference=diff,
                status=status,
                reason=request.form.get('reason', '').strip() or None,
                notes=request.form.get('notes', '').strip() or None,
                counted_by=current_user.id,
                count_date=_parse_date('count_date'),
            )
            db.session.add(row)
            db.session.commit()
            flash('تم تسجيل الجرد السنوي بنجاح.', 'success')
            return redirect(url_for('inventory.counts'))
        except SQLAlchemyError:
            db.session.rollback()
            flash('تعذر حفظ سجل الجرد بسبب خطأ في قاعدة البيانات. يرجى المحاولة مرة أخرى.', 'danger')
        except Exception:
            db.session.rollback()
            flash('تعذر حفظ سجل الجرد. يرجى مراجعة البيانات والمحاولة مرة أخرى.', 'danger')
        return render_template('inventory/count_form.html', items=items, warehouses=warehouses), 400

    return render_template('inventory/count_form.html', items=items, warehouses=warehouses)


@inventory_bp.route('/reports')
@login_required
@any_permission_required('view_inventory', 'manage_inventory')
def reports():
    school, year, response = _school_year_or_redirect()
    if response:
        return response
    items = _items_query(school, year).order_by(InventoryItem.name).all()
    low_items = [item for item in items if item.is_low_stock]
    category_totals = (
        db.session.query(InventoryCategory.name, func.count(InventoryItem.id), func.sum(InventoryItem.current_quantity))
        .join(InventoryItem, InventoryItem.category_id == InventoryCategory.id)
        .filter(InventoryItem.school_id == school.id, InventoryItem.academic_year_id == year.id)
        .group_by(InventoryCategory.name)
        .order_by(InventoryCategory.name)
        .all()
    )
    warehouse_totals = (
        db.session.query(InventoryWarehouse.name,
                         func.count(InventoryItemStock.id),
                         func.sum(InventoryItemStock.quantity))
        .join(InventoryItemStock, InventoryItemStock.warehouse_id == InventoryWarehouse.id)
        .filter(InventoryWarehouse.school_id == school.id,
                InventoryItemStock.academic_year_id == year.id)
        .group_by(InventoryWarehouse.name)
        .order_by(InventoryWarehouse.name)
        .all()
    )
    low_stock_by_warehouse = (
        InventoryItemStock.query
        .options(joinedload(InventoryItemStock.item), joinedload(InventoryItemStock.warehouse))
        .filter(InventoryItemStock.school_id == school.id,
                InventoryItemStock.academic_year_id == year.id,
                InventoryItemStock.quantity <= InventoryItemStock.minimum_quantity)
        .all()
    )
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    movement_query = InventoryMovement.query.filter_by(school_id=school.id, academic_year_id=year.id)
    if from_date:
        movement_query = movement_query.filter(InventoryMovement.movement_date >= datetime.strptime(from_date, '%Y-%m-%d').date())
    if to_date:
        movement_query = movement_query.filter(InventoryMovement.movement_date <= datetime.strptime(to_date, '%Y-%m-%d').date())
    movements_report = movement_query.order_by(InventoryMovement.movement_date.desc()).limit(100).all()
    counts_report = InventoryCount.query.filter_by(school_id=school.id, academic_year_id=year.id).order_by(InventoryCount.count_date.desc()).all()
    return render_template('inventory/reports.html',
                           items=items,
                           low_items=low_items,
                           category_totals=category_totals,
                           warehouse_totals=warehouse_totals,
                           low_stock_by_warehouse=low_stock_by_warehouse,
                           movements=movements_report,
                           counts=counts_report,
                           movement_types=MOVEMENT_TYPES,
                           count_status=COUNT_STATUS,
                           from_date=from_date,
                           to_date=to_date)


@inventory_bp.route('/reports/export.xlsx')
@login_required
@any_permission_required('view_inventory', 'manage_inventory')
def export_excel():
    school, year, response = _school_year_or_redirect()
    if response:
        return response
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception:
        flash('مكتبة Excel غير متاحة.', 'warning')
        return redirect(url_for('inventory.reports'))

    items = _items_query(school, year).order_by(InventoryItem.name).all()
    stocks_by_item = {}
    for stock in (InventoryItemStock.query
                  .options(joinedload(InventoryItemStock.warehouse))
                  .filter(InventoryItemStock.school_id == school.id,
                          InventoryItemStock.academic_year_id == year.id)
                  .all()):
        stocks_by_item.setdefault(stock.item_id, []).append(stock)

    wb = Workbook()
    ws = wb.active
    ws.title = 'المخزون الحالي'
    headers = ['المادة', 'التصنيف', 'الرمز', 'الوحدة', 'الكمية الحالية', 'الحد الأدنى', 'المخازن', 'تنبيه']
    ws.append(headers)
    for item in items:
        warehouse_summary = ' | '.join(
            f'{s.warehouse.name}: {float(s.quantity or 0)}'
            for s in stocks_by_item.get(item.id, []) if s.warehouse
        ) or '-'
        ws.append([
            item.name,
            item.category.name if item.category else '',
            item.item_code or '',
            item.unit,
            float(item.current_quantity or 0),
            float(item.minimum_quantity or 0),
            warehouse_summary,
            'منخفض' if item.is_low_stock else '',
        ])
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1A3A5C')
        cell.alignment = Alignment(horizontal='center')

    ws2 = wb.create_sheet('التوزيع حسب المخزن')
    ws2.append(['المادة', 'التصنيف', 'المخزن', 'الكمية', 'الحد الأدنى', 'تنبيه'])
    for item in items:
        for stock in stocks_by_item.get(item.id, []):
            ws2.append([
                item.name,
                item.category.name if item.category else '',
                stock.warehouse.name if stock.warehouse else '',
                float(stock.quantity or 0),
                float(stock.minimum_quantity or 0),
                'منخفض' if stock.is_low_stock else '',
            ])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1A3A5C')
        cell.alignment = Alignment(horizontal='center')

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=inventory.xlsx'},
    )


@inventory_bp.route('/reports/export.pdf')
@login_required
@any_permission_required('view_inventory', 'manage_inventory')
def export_pdf():
    school, year, response = _school_year_or_redirect()
    if response:
        return response
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from app.utils.pdf_gen import _register_arabic_fonts, _shape_arabic_text, generate_error_pdf
    except Exception:
        flash('مكتبة PDF غير متاحة.', 'warning')
        return redirect(url_for('inventory.reports'))

    arabic_font_registered = _register_arabic_fonts(pdfmetrics, TTFont)
    if not arabic_font_registered:
        pdf_bytes = generate_error_pdf(
            'خطأ في تحميل الخط العربي',
            'يرجى التحقق من وجود ملفات الخط العربي في مجلد static/fonts/'
        )
        if not pdf_bytes:
            flash('تعذر تحميل الخط العربي لتوليد PDF.', 'danger')
            return redirect(url_for('inventory.reports'))
        return Response(pdf_bytes, mimetype='application/pdf')

    title_style = ParagraphStyle(
        'inventory_title',
        fontName='Amiri-Bold',
        fontSize=16,
        alignment=1,
        textColor=colors.HexColor('#1a3a5c'),
        spaceAfter=12,
    )
    header_style = ParagraphStyle(
        'inventory_header',
        fontName='Amiri-Bold',
        fontSize=10,
        alignment=1,
        textColor=colors.white,
    )
    cell_style = ParagraphStyle(
        'inventory_cell',
        fontName='Amiri',
        fontSize=9,
        alignment=1,
        leading=12,
    )

    def ar(text, style=cell_style):
        return Paragraph(_shape_arabic_text(text), style)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    rows = [[
        ar('اسم المادة', header_style),
        ar('التصنيف', header_style),
        ar('الكمية', header_style),
        ar('الحد الأدنى', header_style),
        ar('القيمة', header_style),
    ]]
    stocks_by_item = {}
    for stock in (InventoryItemStock.query
                  .options(joinedload(InventoryItemStock.warehouse))
                  .filter(InventoryItemStock.school_id == school.id,
                          InventoryItemStock.academic_year_id == year.id)
                  .all()):
        stocks_by_item.setdefault(stock.item_id, []).append(stock)

    rows = [['المادة', 'التصنيف', 'الكمية', 'الحد الأدنى', 'المخزن']]
    for item in _items_query(school, year).order_by(InventoryItem.name).all():
        quantity = item.current_quantity or 0
        price = item.purchase_price or 0
        value = quantity * price
        warehouse_summary = ' | '.join(
            f'{s.warehouse.name}: {float(s.quantity or 0)}'
            for s in stocks_by_item.get(item.id, []) if s.warehouse
        ) or '-'
        rows.append([
            ar(item.name),
            ar(item.category.name if item.category else '-'),
            ar(f'{quantity} {item.unit or ""}'),
            ar(str(item.minimum_quantity or 0)),
            ar(str(value)),
            ar(warehouse_summary),
        ])
    rows[0] = [
        ar('اسم المادة', header_style),
        ar('التصنيف', header_style),
        ar('الكمية', header_style),
        ar('الحد الأدنى', header_style),
        ar('القيمة', header_style),
        ar('المخزن', header_style),
    ]
    table = Table(rows, repeatRows=1, colWidths=[110, 85, 70, 65, 65, 100])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3a5c')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Amiri-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Amiri'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7f9fb')]),
    ]))
    doc.build([ar('\u062a\u0642\u0631\u064a\u0631 \u0627\u0644\u0645\u062e\u0632\u0648\u0646', title_style), Spacer(1, 12), table])
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name='inventory.pdf', mimetype='application/pdf')
    doc.build([ar('تقرير المخزون', title_style), Spacer(1, 12), table])
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name='inventory.pdf', mimetype='application/pdf')
    doc.build([Paragraph('تقرير المخزون الحالي', styles['Title']), Spacer(1, 12), table])
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name='inventory.pdf', mimetype='application/pdf')


@inventory_bp.route('/reports/export_annual.xlsx')
@login_required
@any_permission_required('view_inventory', 'manage_inventory')
def export_annual_excel():
    school, year, response = _school_year_or_redirect()
    if response:
        return response
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception:
        flash('مكتبة Excel غير متاحة.', 'warning')
        return redirect(url_for('inventory.reports'))

    counts = (InventoryCount.query
              .options(joinedload(InventoryCount.item).joinedload(InventoryItem.category),
                       joinedload(InventoryCount.warehouse))
              .filter_by(school_id=school.id, academic_year_id=year.id)
              .order_by(InventoryCount.count_date.asc())
              .all())

    wb = Workbook()
    ws = wb.active
    ws.title = 'الجرد السنوي'
    ws.sheet_view.rightToLeft = True

    ws.append([school.school_name if school else '', '', f'السنة: {year.name if year else ""}'])
    ws.append(['الجرد السنوي', '', f'تاريخ التقرير: {date.today().strftime("%Y-%m-%d")}'])
    ws.append([])

    headers = ['التاريخ', 'المادة', 'التصنيف', 'المخزن', 'الحجم', 'كمية النظام', 'الكمية الفعلية', 'الفرق', 'الحالة']
    ws.append(headers)

    status_map = {'match': 'مطابق', 'shortage': 'نقص', 'surplus': 'زيادة'}
    for row in counts:
        item = row.item
        ws.append([
            row.count_date.strftime('%Y-%m-%d') if row.count_date else '',
            item.name if item else '',
            item.category.name if item and item.category else '',
            row.warehouse.name if row.warehouse else '',
            item.size or '' if item else '',
            float(row.system_quantity or 0),
            float(row.actual_quantity or 0),
            float(row.difference or 0),
            status_map.get(row.status, row.status or ''),
        ])

    for cell in ws[4]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1A3A5C')
        cell.alignment = Alignment(horizontal='center')

    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=annual_inventory.xlsx'},
    )


@inventory_bp.route('/reports/export_annual.pdf')
@login_required
@any_permission_required('view_inventory', 'manage_inventory')
def export_annual_pdf():
    school, year, response = _school_year_or_redirect()
    if response:
        return response
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from app.utils.pdf_gen import _register_arabic_fonts, _shape_arabic_text, generate_error_pdf
    except Exception:
        flash('مكتبة PDF غير متاحة.', 'warning')
        return redirect(url_for('inventory.reports'))

    arabic_font_registered = _register_arabic_fonts(pdfmetrics, TTFont)
    if not arabic_font_registered:
        pdf_bytes = generate_error_pdf(
            'خطأ في تحميل الخط العربي',
            'يرجى التحقق من وجود ملفات الخط العربي في مجلد static/fonts/'
        )
        if not pdf_bytes:
            flash('تعذر تحميل الخط العربي لتوليد PDF.', 'danger')
            return redirect(url_for('inventory.reports'))
        return Response(pdf_bytes, mimetype='application/pdf')

    title_style = ParagraphStyle('annual_title', fontName='Amiri-Bold', fontSize=16,
                                 alignment=1, textColor=colors.HexColor('#1a3a5c'), spaceAfter=6)
    sub_style = ParagraphStyle('annual_sub', fontName='Amiri', fontSize=10,
                               alignment=1, textColor=colors.HexColor('#555555'), spaceAfter=12)
    header_style = ParagraphStyle('annual_hdr', fontName='Amiri-Bold', fontSize=9,
                                  alignment=1, textColor=colors.white)
    cell_style = ParagraphStyle('annual_cell', fontName='Amiri', fontSize=8,
                                alignment=1, leading=11)

    def ar(text, style=cell_style):
        return Paragraph(_shape_arabic_text(str(text)), style)

    counts = (InventoryCount.query
              .options(joinedload(InventoryCount.item).joinedload(InventoryItem.category),
                       joinedload(InventoryCount.warehouse))
              .filter_by(school_id=school.id, academic_year_id=year.id)
              .order_by(InventoryCount.count_date.asc())
              .all())

    status_map = {'match': 'مطابق', 'shortage': 'نقص', 'surplus': 'زيادة'}
    table_data = [[
        ar('التاريخ', header_style), ar('المادة', header_style),
        ar('التصنيف', header_style), ar('المخزن', header_style), ar('الحجم', header_style),
        ar('كمية النظام', header_style), ar('الكمية الفعلية', header_style),
        ar('الفرق', header_style), ar('الحالة', header_style),
    ]]
    for row in counts:
        item = row.item
        table_data.append([
            ar(row.count_date.strftime('%Y-%m-%d') if row.count_date else '-'),
            ar(item.name if item else '-'),
            ar(item.category.name if item and item.category else '-'),
            ar(row.warehouse.name if row.warehouse else '-'),
            ar(item.size if item and item.size else '-'),
            ar(str(row.system_quantity or 0)),
            ar(str(row.actual_quantity or 0)),
            ar(str(row.difference or 0)),
            ar(status_map.get(row.status, row.status or '-')),
        ])

    col_widths = [60, 90, 70, 70, 40, 60, 60, 40, 50]
    table = Table(table_data, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3a5c')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7f9fb')]),
    ]))

    school_name = school.school_name if school else ''
    year_name = year.name if year else ''
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            rightMargin=20, leftMargin=20, topMargin=25, bottomMargin=25)
    doc.build([
        ar(school_name, title_style),
        ar(f'الجرد السنوي — {year_name}', title_style),
        Spacer(1, 4),
        ar(f'تاريخ التقرير: {date.today().strftime("%Y-%m-%d")}', sub_style),
        table,
    ])
    buffer.seek(0)
    return send_file(buffer, as_attachment=True,
                     download_name='annual_inventory.pdf', mimetype='application/pdf')
